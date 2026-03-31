"""
Load CardPointe authorization Excel files into BigQuery.
Creates `tds_sales.cardpointe_authorizations` table and MERGEs transaction data.

Usage: python3 scripts/load_cardpointe_authorizations.py [--dry-run]
"""

import os
import sys
import glob
from datetime import datetime
from google.cloud import bigquery
import openpyxl

DOWNLOADS_DIR = os.path.expanduser(
    "~/Downloads/cardpointe reports-1 year of revenue and refunds"
)
CREDENTIALS = os.path.expanduser("~/credentials/bigquery-service-account.json")
PROJECT_ID = "jovial-root-443516-a7"
DATASET = "tds_sales"
TABLE = "cardpointe_authorizations"
FULL_TABLE = f"{PROJECT_ID}.{DATASET}.{TABLE}"

CREATE_TABLE_SQL = f"""
CREATE TABLE IF NOT EXISTS `{FULL_TABLE}` (
  transaction_id STRING NOT NULL,
  location STRING,
  auth_date TIMESTAMP,
  method STRING,
  cardholder_name STRING,
  card_brand STRING,
  last_four STRING,
  amount NUMERIC,
  auth_code STRING,
  status STRING,
  source_file STRING,
  loaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
)
"""

# Only load transactions that represent real money movement
VALID_STATUSES = {"PROCESSED", "REFUNDED", "PARTIALLY_REFUNDED"}
SKIP_METHODS = {"VERIFY"}


def parse_auth_files():
    """Parse all authorization xlsx files and return transaction rows."""
    pattern = os.path.join(DOWNLOADS_DIR, "Authorizations *.xlsx")
    files = sorted(glob.glob(pattern))

    if not files:
        print(f"No authorization files found in {DOWNLOADS_DIR}")
        return []

    print(f"Found {len(files)} authorization files")
    all_rows = []
    skipped = {"voided": 0, "declined": 0, "verify": 0, "other_status": 0}

    for filepath in files:
        filename = os.path.basename(filepath)
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active

        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        col_idx = {h: i for i, h in enumerate(headers)}
        file_count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            txn_id = row[col_idx["Transaction #"]]
            if not txn_id:
                continue

            method = str(row[col_idx["Method"]]).upper().strip()
            status = str(row[col_idx["Status"]]).upper().strip()

            # Skip non-settling transactions
            if method in SKIP_METHODS:
                skipped["verify"] += 1
                continue
            if status == "VOIDED":
                skipped["voided"] += 1
                continue
            if status == "DECLINED":
                skipped["declined"] += 1
                continue
            if status not in VALID_STATUSES:
                skipped["other_status"] += 1
                continue

            # Parse date: "MM/DD/YYYY HH:MM:SS AM/PM"
            date_str = str(row[col_idx["Date"]])
            try:
                auth_date = datetime.strptime(date_str, "%m/%d/%Y %I:%M:%S %p")
            except ValueError:
                try:
                    auth_date = datetime.strptime(date_str, "%m/%d/%Y %H:%M:%S %p")
                except ValueError:
                    print(f"  Warning: Could not parse date '{date_str}' in {filename}")
                    continue

            amount = float(row[col_idx["Amount"]] or 0)

            all_rows.append({
                "transaction_id": str(txn_id).strip(),
                "location": str(row[col_idx["Location"]] or "").strip() or None,
                "auth_date": auth_date.isoformat(),
                "method": method,
                "cardholder_name": str(row[col_idx["Name"]] or "").strip() or None,
                "card_brand": str(row[col_idx["Brand"]] or "").strip() or None,
                "last_four": str(row[col_idx["Last 4"]] or "").strip() or None,
                "amount": amount,
                "auth_code": str(row[col_idx["Auth Code"]] or "").strip() or None,
                "status": status,
                "source_file": filename,
            })
            file_count += 1

        wb.close()
        print(f"  {filename}: {file_count} valid transactions")

    print(f"\nTotal valid: {len(all_rows)}")
    print(f"Skipped - voided: {skipped['voided']}, declined: {skipped['declined']}, "
          f"verify: {skipped['verify']}, other: {skipped['other_status']}")

    return all_rows


def load_to_bigquery(rows, dry_run=False):
    """Create table and MERGE rows into BigQuery."""
    os.environ.setdefault("GOOGLE_APPLICATION_CREDENTIALS", CREDENTIALS)
    client = bigquery.Client(project=PROJECT_ID)

    if dry_run:
        print(f"\n[DRY RUN] Would create table {FULL_TABLE}")
        print(f"[DRY RUN] Would load {len(rows)} rows")
        # Show sample
        for r in rows[:3]:
            print(f"  {r['transaction_id']} | {r['auth_date']} | {r['method']} | "
                  f"${r['amount']:.2f} | {r['status']}")
        return

    # Create table if needed
    print(f"\nCreating table {FULL_TABLE} (if not exists)...")
    client.query(CREATE_TABLE_SQL).result()

    # Load via temp table + MERGE for deduplication
    temp_table = f"{FULL_TABLE}_staging"
    print(f"Loading {len(rows)} rows to staging table...")

    # Define schema for staging
    schema = [
        bigquery.SchemaField("transaction_id", "STRING", mode="REQUIRED"),
        bigquery.SchemaField("location", "STRING"),
        bigquery.SchemaField("auth_date", "TIMESTAMP"),
        bigquery.SchemaField("method", "STRING"),
        bigquery.SchemaField("cardholder_name", "STRING"),
        bigquery.SchemaField("card_brand", "STRING"),
        bigquery.SchemaField("last_four", "STRING"),
        bigquery.SchemaField("amount", "NUMERIC"),
        bigquery.SchemaField("auth_code", "STRING"),
        bigquery.SchemaField("status", "STRING"),
        bigquery.SchemaField("source_file", "STRING"),
    ]

    job_config = bigquery.LoadJobConfig(
        schema=schema,
        write_disposition=bigquery.WriteDisposition.WRITE_TRUNCATE,
    )

    # Convert to dicts BigQuery expects
    load_job = client.load_table_from_json(rows, temp_table, job_config=job_config)
    load_job.result()
    print(f"Staging table loaded: {load_job.output_rows} rows")

    # MERGE into target
    merge_sql = f"""
    MERGE `{FULL_TABLE}` AS target
    USING `{temp_table}` AS source
    ON target.transaction_id = source.transaction_id
    WHEN NOT MATCHED THEN
      INSERT (transaction_id, location, auth_date, method, cardholder_name,
              card_brand, last_four, amount, auth_code, status, source_file, loaded_at)
      VALUES (source.transaction_id, source.location, source.auth_date, source.method,
              source.cardholder_name, source.card_brand, source.last_four, source.amount,
              source.auth_code, source.status, source.source_file, CURRENT_TIMESTAMP())
    """

    print("Merging into target table...")
    merge_result = client.query(merge_sql).result()
    print(f"Merge complete. DML stats: {merge_result.num_dml_affected_rows} rows affected")

    # Clean up staging
    client.delete_table(temp_table, not_found_ok=True)
    print("Staging table cleaned up.")

    # Verify
    verify = client.query(
        f"SELECT COUNT(*) as cnt, MIN(auth_date) as earliest, MAX(auth_date) as latest "
        f"FROM `{FULL_TABLE}`"
    ).result()
    for row in verify:
        print(f"\nVerification: {row.cnt} total rows, "
              f"range {row.earliest} to {row.latest}")


def main():
    dry_run = "--dry-run" in sys.argv

    print("=" * 70)
    print("CardPointe Authorization Data Loader")
    if dry_run:
        print("[DRY RUN MODE]")
    print("=" * 70)

    rows = parse_auth_files()
    if not rows:
        return

    load_to_bigquery(rows, dry_run=dry_run)
    print("\nDone.")


if __name__ == "__main__":
    main()
