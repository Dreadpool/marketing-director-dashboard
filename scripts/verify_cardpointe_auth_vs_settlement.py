"""
Parse CardPointe authorization Excel files and show monthly totals.
Compare against BigQuery settlement data where months overlap.

Usage: python3 scripts/verify_cardpointe_auth_vs_settlement.py
"""

import os
import glob
from collections import defaultdict
from google.cloud import bigquery
import openpyxl

DOWNLOADS_DIR = os.path.expanduser(
    "~/Downloads/cardpointe reports-1 year of revenue and refunds"
)
CREDENTIALS = os.path.expanduser("~/credentials/bigquery-service-account.json")
PROJECT_ID = "jovial-root-443516-a7"
DATASET = "tds_sales"

# Statuses that indicate the transaction will settle
SETTLE_STATUSES = {"PROCESSED", "REFUNDED"}
# Methods that represent real money movement
CHARGE_METHODS = {"SALE", "FORCE"}
REFUND_METHODS = {"REFUND"}
# Skip these entirely
SKIP_STATUSES = {"VOIDED", "DECLINED"}
SKIP_METHODS = {"VERIFY"}


def parse_auth_files():
    """Parse all authorization xlsx files and return monthly aggregates."""
    pattern = os.path.join(DOWNLOADS_DIR, "Authorizations *.xlsx")
    files = sorted(glob.glob(pattern))

    if not files:
        print(f"No authorization files found in {DOWNLOADS_DIR}")
        return {}

    print(f"Found {len(files)} authorization files\n")

    monthly = defaultdict(lambda: {
        "charges": 0.0,
        "refunds": 0.0,
        "charge_count": 0,
        "refund_count": 0,
        "skipped": 0,
        "file": "",
    })

    for filepath in files:
        filename = os.path.basename(filepath)
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active

        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        col_idx = {h: i for i, h in enumerate(headers)}

        for row in ws.iter_rows(min_row=2, values_only=True):
            txn_id = row[col_idx["Transaction #"]]
            if not txn_id:
                continue

            method = str(row[col_idx["Method"]]).upper().strip()
            status = str(row[col_idx["Status"]]).upper().strip()
            amount = float(row[col_idx["Amount"]] or 0)
            date_str = str(row[col_idx["Date"]])

            # Skip transactions that won't settle
            if status in SKIP_STATUSES or method in SKIP_METHODS:
                # Still count for reporting
                month_parts = date_str.split("/")
                if len(month_parts) >= 2:
                    key = f"2025-{int(month_parts[0]):02d}"
                    monthly[key]["skipped"] += 1
                continue

            # Only count transactions with settleable statuses
            if status not in SETTLE_STATUSES:
                continue

            # Extract month from date (MM/DD/YYYY format)
            month_parts = date_str.split("/")
            if len(month_parts) < 2:
                continue
            key = f"2025-{int(month_parts[0]):02d}"
            monthly[key]["file"] = filename

            if method in CHARGE_METHODS:
                monthly[key]["charges"] += amount
                monthly[key]["charge_count"] += 1
            elif method in REFUND_METHODS:
                monthly[key]["refunds"] += abs(amount)
                monthly[key]["refund_count"] += 1

        wb.close()

    return dict(monthly)


def get_settlement_data():
    """Fetch settlement monthly totals from BigQuery."""
    os.environ.setdefault("GOOGLE_APPLICATION_CREDENTIALS", CREDENTIALS)
    client = bigquery.Client(project=PROJECT_ID)

    query = f"""
    SELECT
      EXTRACT(YEAR FROM settlement_date) as yr,
      EXTRACT(MONTH FROM settlement_date) as mo,
      ROUND(SUM(charge_total), 2) as charges,
      ROUND(SUM(refund_total), 2) as refunds,
      ROUND(SUM(net_total), 2) as net,
      COUNT(*) as row_count
    FROM {DATASET}.cardpointe_settlements
    GROUP BY 1, 2
    ORDER BY 1, 2
    """

    results = client.query(query).result()
    settlements = {}
    for row in results:
        key = f"{row.yr}-{row.mo:02d}"
        settlements[key] = {
            "charges": float(row.charges),
            "refunds": float(row.refunds),
            "net": float(row.net),
            "row_count": row.row_count,
        }
    return settlements


def main():
    print("=" * 90)
    print("CardPointe Authorization vs Settlement Verification")
    print("=" * 90)

    auth_data = parse_auth_files()
    if not auth_data:
        return

    settlement_data = get_settlement_data()

    # Print authorization monthly summary
    print("\n--- Authorization Data (from Excel files) ---\n")
    print(
        f"{'Month':<10} {'Charges':>14} {'Refunds':>14} {'Net':>14} "
        f"{'Txns':>8} {'Skipped':>8}"
    )
    print("-" * 78)

    auth_total_charges = 0
    auth_total_refunds = 0
    for key in sorted(auth_data.keys()):
        d = auth_data[key]
        net = d["charges"] - d["refunds"]
        txns = d["charge_count"] + d["refund_count"]
        auth_total_charges += d["charges"]
        auth_total_refunds += d["refunds"]
        print(
            f"{key:<10} ${d['charges']:>12,.2f} ${d['refunds']:>12,.2f} "
            f"${net:>12,.2f} {txns:>8} {d['skipped']:>8}"
        )

    auth_total_net = auth_total_charges - auth_total_refunds
    print("-" * 78)
    print(
        f"{'TOTAL':<10} ${auth_total_charges:>12,.2f} ${auth_total_refunds:>12,.2f} "
        f"${auth_total_net:>12,.2f}"
    )

    # Print settlement monthly summary
    print("\n--- Settlement Data (from BigQuery) ---\n")
    print(f"{'Month':<10} {'Charges':>14} {'Refunds':>14} {'Net':>14} {'Days':>8}")
    print("-" * 66)

    for key in sorted(settlement_data.keys()):
        d = settlement_data[key]
        print(
            f"{key:<10} ${d['charges']:>12,.2f} ${d['refunds']:>12,.2f} "
            f"${d['net']:>12,.2f} {d['row_count']:>8}"
        )

    # Check for overlapping months
    overlap = set(auth_data.keys()) & set(settlement_data.keys())
    if overlap:
        print("\n--- Overlapping Month Comparison ---\n")
        print(
            f"{'Month':<10} {'Auth Charges':>14} {'Settle Charges':>16} "
            f"{'Diff':>12} {'Diff %':>8}"
        )
        print("-" * 66)
        for key in sorted(overlap):
            a = auth_data[key]["charges"]
            s = settlement_data[key]["charges"]
            diff = a - s
            pct = (diff / s * 100) if s > 0 else 0
            print(
                f"{key:<10} ${a:>12,.2f} ${s:>14,.2f} "
                f"${diff:>10,.2f} {pct:>7.2f}%"
            )
    else:
        print("\n--- No overlapping months between auth files and settlement data ---")
        print("Auth files cover: March-December 2025")
        print(f"Settlement data covers: {', '.join(sorted(settlement_data.keys()))}")
        print(
            "\nSince there's no overlap, we cannot directly verify auth vs settlement "
            "accuracy."
        )
        print(
            "Authorization data will be used as the best available proxy for 2025 CC "
            "revenue."
        )


if __name__ == "__main__":
    main()
